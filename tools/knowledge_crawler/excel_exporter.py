"""
Excel导出模块 - 将知识条目导出为Excel文件，匹配 knowledge_base 表结构
"""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import (
    OUTPUT_EXCEL_PATH, OUTPUT_EXCEL_SHEET, EXCEL_COLUMNS, EXCEL_HEADERS,
)


def export_to_excel(knowledge_entries: list[dict], output_path: str = "") -> str:
    if not knowledge_entries:
        print("[警告] 没有知识条目需要导出")
        return ""

    if output_path:
        file_path = output_path
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = OUTPUT_EXCEL_PATH.rsplit(".", 1)[0]
        ext = OUTPUT_EXCEL_PATH.rsplit(".", 1)[1] if "." in OUTPUT_EXCEL_PATH else "xlsx"
        file_path = f"{base_name}_{timestamp}.{ext}"
    
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_EXCEL_SHEET

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, entry in enumerate(knowledge_entries, 2):
        values = [
            entry.get("question", ""),
            entry.get("answer", ""),
            entry.get("category", "通用"),
            entry.get("keywords", ""),
            entry.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    col_widths = [45, 80, 15, 35, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(file_path)
    print(f"\nExcel文件已保存至: {file_path}")
    print(f"共导出 {len(knowledge_entries)} 条知识记录")
    return file_path


def preview_entries(entries: list[dict], max_rows: int = 10):
    if not entries:
        return
    print(f"\n知识条目预览（前 {min(max_rows, len(entries))} 条）：")
    print("-" * 60)
    for i, entry in enumerate(entries[:max_rows]):
        print(f"\n[{i + 1}] 问题: {entry['question']}")
        print(f"    分类: {entry.get('category', '通用')}")
        print(f"    关键词: {entry.get('keywords', '')}")
        print(f"    答案: {entry['answer'][:100]}{'...' if len(entry['answer']) > 100 else ''}")
    print("-" * 60)
